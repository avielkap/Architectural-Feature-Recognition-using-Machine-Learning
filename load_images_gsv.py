# requests and json are the dependencies

import openpyxl
import os

import google_streetview.api
import background_files.addr_proccessing as addr_proccessing
import parameters


def getAddr(addr, dir):
    # geocode_result = gmaps.geocode(addr)
    # for i in geocode_result[0].keys():
    #     print(i)
    #     print(geocode_result[0][i])

    params_format = {
        'size': '500x300',  # max 640x640 pixels
        # 'location': str(geocode_result[0]['geometry']['location']['lat']) + ',' + str(
        #   geocode_result[0]['geometry']['location']['lng']),
        'location': addr,
        # 'heading': '0', #151.78
        'pitch': '30',
        'key': '', # google api id
        'return_error_code': True,
        'radius': 5  # meters
    }
    params = []
    for i in range(6):
        pitch = (i + 1) * 10
        params.append(params_format.copy())
        params[i]['pitch'] = pitch
    results = google_streetview.api.results(params)
    flag = True
    for i in range(parameters.photos_per_addr):
        if results.metadata[i]['status'] != 'OK':
            flag = False
    if flag:
        results.download_links(dir + '/' + addr)
        print(addr + ' downloaded')
    else:
        print(addr + " didn't found")


# for checking
# addr='יהודה הלוי 10 תל אביב יפו'
# getAddr(addr)


def runOverExel(path, destination):
    # Give the location of the file

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

    # Getting the value of maximum rows
    # and column
    row = sheet_obj.max_row
    column = sheet_obj.max_column

    print("Total Rows:", row)
    print("Total Columns:", column)

    # printing the value of first column
    # Loop will print all values
    # of first column
    print('the addresses are on the second column')
    addresses = []
    neighborhoods = []
    for i in range(1, row + 1):
        neighborhoods.append(sheet_obj.cell(row=i, column=1).value)
        address = sheet_obj.cell(row=i, column=2)
        if address.value is not None and any(char.isdigit() for char in address.value):
            addresses.append(address.value)
    full_addresses = addr_proccessing.process_addresses(zip(neighborhoods, addresses))
    neighborhoods, addresses = zip(*full_addresses)
    for addr in addresses:
        addr = addr_proccessing.clear_special_chars(addr)
        if not os.path.exists(destination + '/' + addr):
            getAddr(addr, destination)
    if len(addresses) == len(neighborhoods):
        return addresses, neighborhoods
    print('there are some incorrect addresses')
    return addresses, None
